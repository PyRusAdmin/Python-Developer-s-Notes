from openpyxl import Workbook

# Создаём новую рабочую книгу
wb = Workbook()

# Получаем активный лист
sheet = wb.active

# Записываем данные в ячейки (способ 1: через координаты)
sheet.cell(row=1, column=1).value = "Имя"
sheet.cell(row=1, column=2).value = "Фамилия"
sheet.cell(row=2, column=1).value = "Иван"
sheet.cell(row=2, column=2).value = "Иванов"

# Записываем данные в ячейки (способ 2: через адрес)
sheet["A3"] = "Пётр"
sheet["B3"] = "Петров"

# Сохраняем файл
wb.save("example.xlsx")
